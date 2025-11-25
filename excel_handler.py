from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter


def write_data_to_excel(hang_muc_list, output_file="output.xlsx"):
    wb = Workbook()
    ws = None
    sheet_name = None

    # -------- Create reusable styles -------- #
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_fill = PatternFill("solid", fgColor="BDD7EE")

    header_style = NamedStyle(name="header_style")
    header_style.font = header_font
    header_style.alignment = center
    header_style.border = border
    header_style.fill = header_fill

    normal_style = NamedStyle(name="normal_style")
    normal_style.alignment = Alignment(wrap_text=True, vertical="top")
    normal_style.border = border

    if "header_style" not in wb.named_styles:
        wb.add_named_style(header_style)
    if "normal_style" not in wb.named_styles:
        wb.add_named_style(normal_style)

    # -------- Iterate HangMuc -------- #
    for hang_muc in hang_muc_list:
        # If has name → new sheet
        if hang_muc.ten_hang_muc.strip() != "":
            sheet_name = hang_muc.ten_hang_muc[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)

            # Write section title on top
            ws.merge_cells("A1:D1")
            ws["A1"].value = hang_muc.ten_hang_muc
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = center

            # Write header row
            headers = ["STT", "Nội dung công việc", "Đơn vị", "Khối lượng"]
            ws.append(headers)

            for col in range(1, 5):
                ws.cell(row=2, column=col).style = "header_style"

            start_row = 3

        # If no name → continues writing into previous sheet
        if ws is None:
            raise ValueError("First HangMuc must have ten_hang_muc (sheet name)")

        # -------- Write rows -------- #
        row_index = ws.max_row
        if row_index < 3:
            row_index = 3

        stt = 1
        for cv in hang_muc.cong_viec:
            ws.append([stt, cv.noi_dung_cong_viec, cv.don_vi, cv.khoi_luong])
            current_row = ws.max_row

            # Apply styles to each cell
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).style = "normal_style"

            # Right-align number
            ws.cell(row=current_row, column=4).number_format = "#,##0.000"
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")

            stt += 1

        # Auto column width
        # for col in ["A", "B", "C", "D"]:
        #     ws.column_dimensions[col].width = 20
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = length + 2

    # Remove default empty sheet if still exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_file)
    print("✔ Exported to", output_file)
