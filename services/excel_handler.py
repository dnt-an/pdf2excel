# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
# from openpyxl.utils import get_column_letter


# def write_data_to_excel(hang_muc_list, output_file="output.xlsx"):
#     wb = Workbook()
#     ws = None
#     sheet_name = None

#     # -------- Create reusable styles -------- #
#     header_font = Font(bold=True)
#     center = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     border = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin"),
#     )

#     header_fill = PatternFill("solid", fgColor="BDD7EE")

#     header_style = NamedStyle(name="header_style")
#     header_style.font = header_font
#     header_style.alignment = center
#     header_style.border = border
#     header_style.fill = header_fill

#     normal_style = NamedStyle(name="normal_style")
#     normal_style.alignment = Alignment(wrap_text=True, vertical="top")
#     normal_style.border = border

#     if "header_style" not in wb.named_styles:
#         wb.add_named_style(header_style)
#     if "normal_style" not in wb.named_styles:
#         wb.add_named_style(normal_style)

#     # -------- Iterate HangMuc -------- #
#     for hang_muc in hang_muc_list:
#         # If has name → new sheet
#         if hang_muc.ten_hang_muc.strip() != "":
#             sheet_name = hang_muc.ten_hang_muc[:31]  # Excel sheet name limit
#             ws = wb.create_sheet(title=sheet_name)

#             # Write section title on top
#             ws.merge_cells("A1:D1")
#             ws["A1"].value = hang_muc.ten_hang_muc
#             ws["A1"].font = Font(bold=True, size=14)
#             ws["A1"].alignment = center

#             # Write header row
#             headers = ["STT", "Nội dung công việc", "Đơn vị", "Khối lượng"]
#             ws.append(headers)

#             for col in range(1, 5):
#                 ws.cell(row=2, column=col).style = "header_style"

#             start_row = 3

#         # If no name → continues writing into previous sheet
#         if ws is None:
#             raise ValueError("First table should have Hang Muc information")

#         # -------- Write rows -------- #
#         row_index = ws.max_row
#         if row_index < 3:
#             row_index = 3

#         stt = 1
#         for cv in hang_muc.cong_viec:
#             ws.append([stt, cv.noi_dung_cong_viec, cv.don_vi, cv.khoi_luong])
#             current_row = ws.max_row

#             # Apply styles to each cell
#             for col in range(1, 5):
#                 ws.cell(row=current_row, column=col).style = "normal_style"

#             # Right-align number
#             ws.cell(row=current_row, column=4).number_format = "#,##0.000"
#             ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")

#             stt += 1

#         # Auto column width
#         # for col in ["A", "B", "C", "D"]:
#         #     ws.column_dimensions[col].width = 20
#         for column_cells in ws.columns:
#             length = max(
#                 len(str(cell.value)) if cell.value is not None else 0
#                 for cell in column_cells
#             )
#             column_letter = get_column_letter(column_cells[0].column)
#             ws.column_dimensions[column_letter].width = length + 2

#     # Remove default empty sheet if still exists
#     if "Sheet" in wb.sheetnames:
#         del wb["Sheet"]

#     wb.save(output_file)
#     print("✔ Exported to", output_file)


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------


def write_data_to_excel(hang_muc_list, output_file="output_incremental_stt.xlsx"):
    wb = Workbook()
    ws = None

    # Define common styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    header_font = Font(bold=True)
    header_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_style = NamedStyle(name="header_style")
    header_style.font = header_font
    header_style.alignment = header_center
    header_style.border = thin_border
    header_style.fill = header_fill

    normal_style = NamedStyle(name="normal_style")
    normal_style.alignment = Alignment(wrap_text=True, vertical="top")
    normal_style.border = thin_border

    number_style = NamedStyle(name="number_style")
    number_style.alignment = Alignment(horizontal="right", vertical="top")
    number_style.border = thin_border
    # number_style.number_format = "#,##0.000"

    unit_style = NamedStyle(name="unit_style")
    unit_style.alignment = Alignment(horizontal="center", vertical="top")
    unit_style.border = thin_border

    # Add styles to workbook
    for style in [header_style, normal_style, number_style, unit_style]:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    NUM_COLUMNS = 5

    # We need a counter to handle the main numbering (1, 2, 3...)
    main_item_counter = 1

    # -------- Iterate HangMuc -------- #
    for hang_muc in hang_muc_list:
        if hang_muc.ten_hang_muc.strip() != "":
            sheet_name = hang_muc.ten_hang_muc[:31].replace(":", "").replace("/", "")
            ws = wb.create_sheet(title=sheet_name)

            start_row = 1

            # 1. Write main item (Hang Muc) - Row 1
            ws.cell(
                row=start_row, column=1
            ).value = f"{main_item_counter}. {hang_muc.ten_hang_muc}"
            ws.merge_cells(
                start_row=start_row,
                end_row=start_row,
                start_column=1,
                end_column=NUM_COLUMNS,
            )

            hang_muc_cell = ws.cell(row=start_row, column=1)
            hang_muc_cell.style = "normal_style"
            hang_muc_cell.font = Font(bold=True)
            hang_muc_cell.alignment = Alignment(vertical="center", wrap_text=True)

            start_row += 1

            # 2. Write header row (Row 2)
            headers_main = [
                "STT (1)",
                "Mô tả công việc mời thầu (2)",
                "Yêu cầu kỹ thuật/ Chỉ dẫn kỹ thuật chính (3)",
                "Khối lượng mời thầu (4)",
                "Đơn vị tính (5)",
            ]
            ws.append(headers_main)

            # Apply header style
            for col in range(1, NUM_COLUMNS + 1):
                ws.cell(row=start_row, column=col).style = "header_style"

            data_start_row = start_row + 1

            # Increment main counter for the next HangMuc
            main_item_counter += 1

        if ws is None:
            # Should only happen if hang_muc_list is empty, but we keep the guardrail
            raise ValueError("First table should have Hang Muc information")

        # -------- Write data rows with incremental STT -------- #

        # We use main_item_counter - 1 because it was just incremented for the next HangMuc
        # This provides the leading number (e.g., '1')
        base_stt = main_item_counter - 1

        for i, cv in enumerate(hang_muc.cong_viec):
            # Calculate the incremental STT: 1.1, 1.2, 1.3...
            sub_stt = f"{base_stt}.{cv.stt}"

            row_data = [
                sub_stt,
                cv.noi_dung_cong_viec,
                "Theo quy định tại Chương V",
                cv.khoi_luong,
                cv.don_vi,
            ]
            ws.append(row_data)
            current_row = ws.max_row

            # Apply specific styles for 4 columns:

            # Col 1 (STT): Normal style, Center aligned
            stt_cell = ws.cell(row=current_row, column=1)
            stt_cell.style = "normal_style"
            stt_cell.alignment = header_center

            # Col 2 (Mô tả công việc): Normal style
            ws.cell(row=current_row, column=2).style = "normal_style"

            # Col 3 (Yêu cầu kỹ thuật): Normal style
            ws.cell(row=current_row, column=3).style = "normal_style"

            # # Col 4 (Khối lượng): Number style (Right-aligned, formatted)
            ws.cell(row=current_row, column=4).style = "number_style"

            # Col 5 (Đơn vị tính): Unit style (Centered)
            ws.cell(row=current_row, column=5).style = "unit_style"

        # Auto-adjust column width
        for i, column_cells in enumerate(ws.columns):
            column_letter = get_column_letter(column_cells[0].column)

            if column_letter == "B":
                ws.column_dimensions[column_letter].width = 60.0
            elif column_letter == "C":
                ws.column_dimensions[column_letter].width = 15.0
            else:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells[1:]
                )
                ws.column_dimensions[column_letter].width = length + 2

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_file)
    print("✔ Exported data to", output_file)
